from __future__ import annotations

import io
import json
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Callable
from zipfile import ZipFile


class LoaderError(ValueError):
    """输入压缩包读取失败。"""


@dataclass
class LoadedEvalLog:
    zip_path: Path
    zip_name: str
    log_name: str
    log_text: str
    xlsx_name: str | None
    xlsx_rows: list[dict[str, Any]] | None
    export_data_list: list[Any] | None
    empty_result: Any | None
    overlength_result: Any | None
    timeout_result: Any | None
    duplicate_result: Any | None
    discovered_files: dict[str, str | None]


def load_eval_zip(zip_path: str | Path) -> LoadedEvalLog:
    """直接从 zip 中读取评测文件，不落盘解压，避免 Windows 长路径问题。"""
    source = Path(zip_path)
    if not source.exists():
        raise FileNotFoundError(f"zip 文件不存在: {source}")
    if not source.is_file():
        raise LoaderError(f"zip 路径不是文件: {source}")

    with ZipFile(source) as zf:
        names = zf.namelist()
        discovered = _discover_names(names)
        log_path = discovered["log"]
        if log_path is None:
            raise LoaderError("zip 中未发现 .log 文件，无法分析 retry 链路")

        log_text = _read_text(zf, log_path)
        xlsx_name = discovered["xlsx"]
        xlsx_rows = read_xlsx_from_zip(zf, xlsx_name) if xlsx_name else None

        def json_or_none(name: str | None) -> Any:
            if name is None:
                return None
            return read_json_flexible_text(_read_text(zf, name))

        return LoadedEvalLog(
            zip_path=source,
            zip_name=source.name,
            log_name=PurePosixPath(log_path).name,
            log_text=log_text,
            xlsx_name=PurePosixPath(xlsx_name).name if xlsx_name else None,
            xlsx_rows=xlsx_rows,
            export_data_list=_as_list(json_or_none(discovered["export_data_list"])),
            empty_result=json_or_none(discovered["empty_result"]),
            overlength_result=json_or_none(discovered["overlength_result"]),
            timeout_result=json_or_none(discovered["timeout"]),
            duplicate_result=json_or_none(discovered["duplicate_result"]),
            discovered_files={k: (PurePosixPath(v).name if v else None) for k, v in discovered.items()},
        )


def _read_text(zf: ZipFile, name: str) -> str:
    return zf.read(name).decode("utf-8", errors="replace")


def read_json_flexible_text(text: str) -> Any:
    """读取普通 JSON、JSONL 或连续 JSON 对象。"""
    text = text.strip()
    if not text:
        return []
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    jsonl_items: list[Any] = []
    jsonl_ok = True
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            jsonl_items.append(json.loads(line))
        except json.JSONDecodeError:
            jsonl_ok = False
            break
    if jsonl_ok and jsonl_items:
        return jsonl_items

    decoder = json.JSONDecoder()
    index = 0
    items = []
    while index < len(text):
        while index < len(text) and text[index].isspace():
            index += 1
        if index >= len(text):
            break
        try:
            item, end = decoder.raw_decode(text, index)
        except json.JSONDecodeError as exc:
            raise LoaderError(f"JSON 文本解析失败: {exc}") from exc
        items.append(item)
        index = end
    return items


def read_json_flexible(path: str | Path) -> Any:
    """从文件路径读取 JSON，兼容多种格式。"""
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    return read_json_flexible_text(text)


def read_xlsx_from_zip(zf: ZipFile, name: str) -> list[dict[str, Any]]:
    """从 zip 内直接读取 xlsx。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise LoaderError("读取 xlsx 需要安装 openpyxl") from exc

    data = zf.read(name)
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return result


def read_xlsx(path: str | Path) -> list[dict[str, Any]]:
    """读取 xlsx 首行表头为字典列表；缺少 openpyxl 时给出明确错误。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise LoaderError("读取 xlsx 需要安装 openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return result


def _discover_names(names: list[str]) -> dict[str, str | None]:
    """根据 zip 内文件名列表发现各类关键文件。"""

    def first_match(predicate: Callable[[str], bool]) -> str | None:
        for name in sorted(names):
            if predicate(name):
                return name
        return None

    return {
        "log": first_match(lambda name: name.endswith(".log")),
        "xlsx": first_match(lambda name: name.endswith(".xlsx")),
        "export_data_list": first_match(
            lambda name: PurePosixPath(name).name.endswith("_export_data_list.json")
        ),
        "empty_result": first_match(lambda name: PurePosixPath(name).name.endswith("_empty_result.json")),
        "overlength_result": first_match(
            lambda name: PurePosixPath(name).name.endswith("_overlength_result.json")
        ),
        "timeout": first_match(lambda name: PurePosixPath(name).name.endswith("_timeout.json")),
        "duplicate_result": first_match(
            lambda name: PurePosixPath(name).name.endswith("_duplicate_result.json")
        ),
    }


def _as_list(value: Any) -> list[Any] | None:
    if value is None:
        return None
    if isinstance(value, list):
        return value
    return [value]
